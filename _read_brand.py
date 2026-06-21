import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\yu\Desktop\耳机品牌统计.xlsx')
ws = wb['蓝牙耳机品牌']
print('COLUMNS:', [ws.cell(1,c).value for c in range(1, ws.max_column+1)])
for r in range(2, ws.max_row+1):
    vals = [str(ws.cell(r,c).value or '') for c in range(1, ws.max_column+1)]
    print('ROW', r, ':', ' | '.join(vals))
