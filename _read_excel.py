import openpyxl, sys, traceback
try:
    wb = openpyxl.load_workbook(r'C:\Users\yu\Desktop\耳机品牌统计.xlsx', data_only=True)
    print('Sheets:', wb.sheetnames)
    ws = wb.active
    print('Active sheet:', ws.title)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 10:
            print(f'Row {i}: {row}')
        for cell in row:
            if cell and 'Newyu' in str(cell):
                print(f'\nFound Newyu at row {i}: {row}')
except Exception as e:
    traceback.print_exc()
    sys.exit(1)
