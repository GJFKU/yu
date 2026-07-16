import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\yu\Desktop\耳机品牌统计.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\n--- {sn} ---')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 15:
            print(f'Row {i}: {list(row)}')
        for cell in row:
            if cell and 'Newyu' in str(cell):
                print(f'*** FOUND Newyu at {sn} row {i}: {list(row)}')
