import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\yu\Desktop\耳机品牌统计.xlsx')
ws = wb['蓝牙耳机品牌']
print(f"Total rows: {ws.max_row}, Total cols: {ws.max_column}")
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
print(f"Headers: {headers}")
# Find 渡哲特
for r in range(2, ws.max_row+1):
    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
    if row[0] and '渡哲' in str(row[0]):
        print(f"Found: {row}")
# Also print all brand names
print("\nAll brands:")
for r in range(2, ws.max_row+1):
    name = ws.cell(row=r, column=1).value
    if name:
        print(f"  {name}")
